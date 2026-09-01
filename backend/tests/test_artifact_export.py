"""Tests for the artifact export pipeline (Task 6 — artifact-export skills).

Covers the four renderer modules (PDF / PPTX / XLSX / CSV) and the
``ExportService`` orchestration layer.  The renderers are tested in
isolation (bytes in / bytes out, no DB) so they can run in any
environment, and the service is tested with a SQLite-in-memory DB
plus a single ``Artifact`` fixture.

The renderers are deliberately small + deterministic — if you change
the visual layout, just update the test expectations rather than the
renderer.  We only assert:

  * Output is non-empty bytes
  * Magic-byte header is correct (PDF starts with %PDF, PPTX/XLSX are
    ZIP archives and start with ``PK\x03\x04``, CSV starts with the
    BOM + # header)
  * Filename / mime-type / extension are sane
  * The service caches the render (a second call doesn't re-render)
  * The service falls back gracefully when the format blob already
    exists on a version
"""

import io
import os
import sys
import unittest
import zipfile
from unittest.mock import MagicMock, patch

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_BACKEND_ROOT = os.path.abspath(os.path.join(_THIS_DIR, ".."))
if _BACKEND_ROOT not in sys.path:
    sys.path.insert(0, _BACKEND_ROOT)
os.chdir(_BACKEND_ROOT)


# ---------------------------------------------------------------------------
# Sample payload — shared by all tests
# ---------------------------------------------------------------------------


def _sample_payload():
    """Build a representative `ReportCardPayload` for the tests."""
    from app.services.synexia.contracts import (
        ReportCardPayload, KPISpec, ChartSpec, InsightSpec, ActionSpec,
    )
    return ReportCardPayload(
        title="Top materials by revenue",
        source="erp_v_sale_orderentry · db_zhanlu_no1",
        generated_at="2026-07-13T08:30:00Z",
        summary=(
            "Top 7 materials account for 76% of revenue; "
            "\u78b3\u4e94\u77f3\u6cb9\u6811\u8102 alone is 35%."
        ),
        kpis=[
            KPISpec(label="Total revenue",  value="189.3M CNY", delta="+12%",
                    caption="Top 7 materials"),
            KPISpec(label="Total quantity", value="11,210 tons", caption="All time"),
            KPISpec(label="Top share",      value="35%", caption="\u78b3\u4e94\u77f3\u6cb9\u6811\u8102"),
            KPISpec(label="Row count",      value="7", caption="Distinct materials"),
        ],
        chart=ChartSpec(
            type="bar",
            title="Top materials by revenue",
            x_key="material_name",
            y_keys=["total_revenue"],
            unit="CNY",
            data=[
                {"material_name": "\u78b3\u4e94\u77f3\u6cb9\u6811\u8102", "total_revenue": 66_255_000},
                {"material_name": "Material B", "total_revenue": 22_100_000},
                {"material_name": "Material C", "total_revenue": 18_700_000},
                {"material_name": "Material D", "total_revenue": 12_500_000},
                {"material_name": "Material E", "total_revenue": 8_300_000},
            ],
        ),
        insights=[
            InsightSpec(icon="trending-up", text="Top 3 materials account for 76% of revenue."),
            InsightSpec(icon="alert-triangle", text="Concentration risk worth monitoring."),
            InsightSpec(icon="lightbulb",     text="\u78b3\u4e94\u77f3\u6cb9\u6811\u8102 is 3x the next material."),
        ],
        next_step="Want to break this down by region, or save this as a weekly recurring report?",
        actions=[
            ActionSpec(label="Break down by region", prompt="Break this down by region."),
            ActionSpec(label="Save as weekly",       prompt="Save this as a recurring weekly report."),
        ],
        user_signal="export",
        warnings=["Snapshot was capped to 5 rows for the chart slide."],
    )


# ---------------------------------------------------------------------------
# Format registry / dispatch
# ---------------------------------------------------------------------------


class TestExportRegistry(unittest.TestCase):
    """The package's public surface — ``render()`` dispatch + helpers."""

    def test_supported_formats(self):
        from app.services.artifacts.exporters import SUPPORTED_FORMATS
        self.assertEqual(
            set(SUPPORTED_FORMATS), {"pdf", "pptx", "xlsx", "csv", "docx", "html"}
        )

    def test_render_unknown_format(self):
        from app.services.artifacts.exporters import render
        data, mime, ext = render("odp", _sample_payload())
        self.assertEqual(data, b"")
        self.assertEqual(mime, "application/octet-stream")
        self.assertEqual(ext, "")

    def test_render_empty_format(self):
        from app.services.artifacts.exporters import render
        data, mime, ext = render("", _sample_payload())
        self.assertEqual(ext, "")

    def test_safe_file_extension(self):
        from app.services.artifacts.exporters import safe_file_extension
        self.assertEqual(safe_file_extension("pdf"),  ".pdf")
        self.assertEqual(safe_file_extension("PDF"),  ".pdf")
        self.assertEqual(safe_file_extension("pptx"), ".pptx")
        self.assertEqual(safe_file_extension("xlsx"), ".xlsx")
        self.assertEqual(safe_file_extension("csv"),  ".csv")
        self.assertEqual(safe_file_extension("docx"), ".docx")
        self.assertEqual(safe_file_extension("odp"),  "")

    def test_safe_mime_type(self):
        from app.services.artifacts.exporters import safe_mime_type
        self.assertIn("pdf", safe_mime_type("pdf"))
        self.assertIn("presentationml", safe_mime_type("pptx"))
        self.assertIn("spreadsheetml", safe_mime_type("xlsx"))
        self.assertIn("csv", safe_mime_type("csv"))


# ---------------------------------------------------------------------------
# PDF renderer
# ---------------------------------------------------------------------------


class TestPDFExport(unittest.TestCase):
    """PDF export — uses reportlab, so we assert on magic bytes + size."""

    def test_render_returns_valid_pdf(self):
        from app.services.artifacts.exporters import render
        data, mime, ext = render("pdf", _sample_payload())
        self.assertTrue(data.startswith(b"%PDF-"), "PDF must start with %PDF- magic")
        self.assertGreater(len(data), 2000, "PDF should not be empty")
        self.assertEqual(mime, "application/pdf")
        self.assertEqual(ext, ".pdf")

    def test_render_handles_unicode_title(self):
        from app.services.artifacts.exporters import render
        p = _sample_payload()
        p.title = "\u62a5\u544a\u6807\u9898 with English"
        data, _, _ = render("pdf", p)
        self.assertTrue(data.startswith(b"%PDF-"))

    def test_render_handles_no_chart(self):
        from app.services.artifacts.exporters import render
        p = _sample_payload()
        p.chart = None
        data, _, _ = render("pdf", p)
        self.assertTrue(data.startswith(b"%PDF-"))

    def test_render_handles_empty_kpis(self):
        from app.services.artifacts.exporters import render
        p = _sample_payload()
        p.kpis = []
        data, _, _ = render("pdf", p)
        self.assertTrue(data.startswith(b"%PDF-"))

    def test_render_handles_line_chart(self):
        from app.services.artifacts.exporters import render
        p = _sample_payload()
        p.chart.type = "line"
        data, _, _ = render("pdf", p)
        self.assertTrue(data.startswith(b"%PDF-"))

    def test_render_handles_pie_chart(self):
        from app.services.artifacts.exporters import render
        p = _sample_payload()
        p.chart.type = "pie"
        data, _, _ = render("pdf", p)
        self.assertTrue(data.startswith(b"%PDF-"))

    def test_render_with_export_context(self):
        from app.services.artifacts.exporters import render, ExportContext
        data, _, _ = render(
            "pdf", _sample_payload(),
            ExportContext(source="erp_v_sale_orderentry", sql="SELECT 1", conversation_id="c-1"),
        )
        self.assertTrue(data.startswith(b"%PDF-"))


# ---------------------------------------------------------------------------
# PPTX renderer
# ---------------------------------------------------------------------------


class TestPPTXExport(unittest.TestCase):
    """PPTX export — uses python-pptx, so we assert on ZIP + slide count."""

    def test_render_returns_valid_pptx(self):
        from app.services.artifacts.exporters import render
        data, mime, ext = render("pptx", _sample_payload())
        # PPTX is a ZIP
        self.assertEqual(data[:4], b"PK\x03\x04", "PPTX must be a ZIP archive")
        self.assertIn("presentationml", mime)
        self.assertEqual(ext, ".pptx")

        with zipfile.ZipFile(io.BytesIO(data)) as z:
            slide_files = [n for n in z.namelist() if n.startswith("ppt/slides/slide")]
            self.assertGreaterEqual(len(slide_files), 4,
                "Expected at least 4 slides (title, summary, kpi, chart)")

    def test_render_minimal_payload(self):
        """A bare-bones payload (no KPIs, no chart) should still produce a valid PPTX
        (at least a title slide + summary slide if summary exists)."""
        from app.services.artifacts.exporters import render
        from app.services.synexia.contracts import ReportCardPayload
        p = ReportCardPayload(title="Minimal", summary="Just a summary.", user_signal="default")
        data, _, _ = render("pptx", p)
        self.assertEqual(data[:4], b"PK\x03\x04")

        with zipfile.ZipFile(io.BytesIO(data)) as z:
            slide_files = [n for n in z.namelist() if n.startswith("ppt/slides/slide")]
            self.assertGreaterEqual(len(slide_files), 1)

    def test_render_handles_no_summary_no_kpis(self):
        from app.services.artifacts.exporters import render
        from app.services.synexia.contracts import ReportCardPayload
        p = ReportCardPayload(title="No metadata", user_signal="default")
        data, _, _ = render("pptx", p)
        self.assertEqual(data[:4], b"PK\x03\x04")


# ---------------------------------------------------------------------------
# XLSX renderer
# ---------------------------------------------------------------------------


class TestXLSXExport(unittest.TestCase):
    """XLSX export — uses openpyxl, re-read the bytes to assert sheets."""

    def test_render_returns_valid_xlsx(self):
        from app.services.artifacts.exporters import render
        data, mime, ext = render("xlsx", _sample_payload())
        self.assertEqual(data[:4], b"PK\x03\x04", "XLSX must be a ZIP archive")
        self.assertIn("spreadsheetml", mime)
        self.assertEqual(ext, ".xlsx")

        # Reload via openpyxl to confirm the workbook is well-formed
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("KPIs", wb.sheetnames)
        self.assertIn("Data", wb.sheetnames)
        self.assertIn("Insights", wb.sheetnames)

    def test_kpi_sheet_has_one_row_per_kpi(self):
        from app.services.artifacts.exporters import render
        from openpyxl import load_workbook
        data, _, _ = render("xlsx", _sample_payload())
        wb = load_workbook(io.BytesIO(data), read_only=True)
        kpi = wb["KPIs"]
        # header + 4 KPI rows
        rows = list(kpi.iter_rows(values_only=True))
        self.assertEqual(len(rows), 5)
        self.assertEqual(rows[0], ("Label", "Value", "Delta", "Caption"))

    def test_data_sheet_has_chart_rows(self):
        from app.services.artifacts.exporters import render
        from openpyxl import load_workbook
        data, _, _ = render("xlsx", _sample_payload())
        wb = load_workbook(io.BytesIO(data), read_only=True)
        data_sheet = wb["Data"]
        rows = list(data_sheet.iter_rows(values_only=True))
        # header (3rd row) + 5 data rows
        data_rows = [r for r in rows if r and r[0] is not None and r[0] != "Data"
                     and r[0] != "Top materials by revenue"]
        # The first non-empty row is the header
        self.assertEqual(data_rows[0], ("material_name", "total_revenue"))
        self.assertEqual(len(data_rows), 6)  # header + 5 rows

    def test_render_minimal_payload(self):
        from app.services.artifacts.exporters import render
        from app.services.synexia.contracts import ReportCardPayload
        p = ReportCardPayload(title="Minimal", summary="Just a summary.", user_signal="default")
        data, _, _ = render("xlsx", p)
        self.assertEqual(data[:4], b"PK\x03\x04")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True)
        self.assertIn("Summary", wb.sheetnames)


# ---------------------------------------------------------------------------
# CSV renderer
# ---------------------------------------------------------------------------


class TestCSVExport(unittest.TestCase):
    """CSV export — pure-stdlib, byte-exact assertions."""

    def test_render_returns_csv_with_bom(self):
        from app.services.artifacts.exporters import render
        data, mime, ext = render("csv", _sample_payload())
        self.assertEqual(data[:3], b"\xef\xbb\xbf", "CSV must start with UTF-8 BOM")
        self.assertIn("csv", mime)
        self.assertEqual(ext, ".csv")
        # Decoded, the first non-empty line should be a # header
        text = data.decode("utf-8-sig")
        lines = [l for l in text.splitlines() if l.strip()]
        self.assertTrue(lines[0].startswith("# Title:"))
        self.assertTrue(any(l.startswith("# user_signal:") for l in lines))

    def test_csv_includes_data_rows(self):
        from app.services.artifacts.exporters import render
        data, _, _ = render("csv", _sample_payload())
        text = data.decode("utf-8-sig")
        self.assertIn("material_name", text)
        self.assertIn("total_revenue", text)
        # The numeric top-row should appear (no decimal, raw int)
        self.assertIn("66255000", text)

    def test_csv_handles_no_data(self):
        from app.services.artifacts.exporters import render
        from app.services.synexia.contracts import ReportCardPayload
        p = ReportCardPayload(title="Empty", user_signal="default")
        data, _, _ = render("csv", p)
        text = data.decode("utf-8-sig")
        self.assertIn("No data in this report.", text)

    def test_csv_one_line_collapses_whitespace(self):
        from app.services.artifacts.exporters.csv_export import _one_line
        self.assertEqual(_one_line("a\nb\nc"), "a b c")
        self.assertEqual(_one_line(""), "")
        long = "x" * 300
        out = _one_line(long)
        self.assertLessEqual(len(out), 250)
        self.assertTrue(out.endswith("\u2026"))


# ---------------------------------------------------------------------------
# Common helpers
# ---------------------------------------------------------------------------


class TestCommonHelpers(unittest.TestCase):
    """The pure-Python helpers in `_common`."""

    def test_slugify_title(self):
        from app.services.artifacts.exporters._common import slugify_title
        self.assertEqual(slugify_title("Sales report — top materials!"),
                         "sales-report-top-materials")
        self.assertEqual(slugify_title(""), "report")
        self.assertEqual(slugify_title("!!!"), "report")
        # CJK title — CJK chars kept as-is
        self.assertEqual(slugify_title("\u62a5\u544a"), "\u62a5\u544a")
        self.assertEqual(slugify_title("Hello 世界!"), "hello-世界")

    def test_stamp_filename_has_timestamp(self):
        from app.services.artifacts.exporters._common import stamp_filename
        fn = stamp_filename("Hello world", "pdf")
        self.assertTrue(fn.startswith("hello-world-"))
        self.assertTrue(fn.endswith(".pdf"))
        # timestamp is YYYYMMDD-HHMM
        import re
        self.assertRegex(fn, r"-\d{8}-\d{4}\.pdf$")

    def test_coerce_number_accepts_strings_with_units(self):
        from app.services.artifacts.exporters._common import coerce_number
        self.assertEqual(coerce_number("1,234.5"), 1234.5)
        self.assertEqual(coerce_number("100%"), 100.0)
        self.assertEqual(coerce_number("$1,000"), 1000.0)
        self.assertEqual(coerce_number("abc"), None)
        self.assertEqual(coerce_number(None), None)
        self.assertEqual(coerce_number(""), None)
        self.assertEqual(coerce_number(42), 42.0)
        self.assertEqual(coerce_number(3.14), 3.14)

    def test_chart_helpers_without_chart(self):
        from app.services.artifacts.exporters._common import (
            chart_rows, chart_x_key, chart_y_keys, chart_x_value,
        )
        from app.services.synexia.contracts import ReportCardPayload
        p = ReportCardPayload(title="t", user_signal="default")
        self.assertEqual(chart_rows(p), [])
        self.assertEqual(chart_x_key(p), "label")
        self.assertEqual(chart_y_keys(p), ["value"])
        self.assertEqual(chart_x_value({}, p), "")

    def test_insight_icon_to_emoji(self):
        from app.services.artifacts.exporters._common import insight_icon_to_emoji
        self.assertIn(insight_icon_to_emoji("lightbulb"), "\U0001F4A1")
        self.assertEqual(insight_icon_to_emoji("unknown"), "\u2022")
        self.assertEqual(insight_icon_to_emoji(""), "\u2022")


# ---------------------------------------------------------------------------
# ExportService — orchestration (uses SQLite-in-memory for the DB)
# ---------------------------------------------------------------------------


def _make_artifact_with_payload(db, payload=None, *, conversation_id: str = "c-test"):
    """Create an Artifact + version + HTML blob directly, then attach
    a ReportCardPayload to metadata_json so ExportService can find it.
    """
    from datetime import datetime
    from uuid import uuid4
    from app.models.artifact import Artifact, ArtifactVersion, ArtifactBlob
    from app.services.artifacts.artifact_service import ArtifactService

    payload = payload or _sample_payload()
    artifact = Artifact(
        id=str(uuid4()),
        conversation_id=conversation_id,
        created_by_agent_id="test-agent",
        artifact_type="html_report",
        title=payload.title,
        description=payload.summary,
        status="preview_ready",
        visibility="conversation_private",
        tags=["report", "test"],
        metadata_json={
            "report_card_payload": payload.model_dump(),
            "source": payload.source,
            "sql": "SELECT 1",
            "user_signal": payload.user_signal,
            "payload_formats": {},
        },
    )
    db.add(artifact)
    db.flush()

    version = ArtifactVersion(
        id=str(uuid4()),
        artifact_id=artifact.id,
        version_number=1,
        status="preview_ready",
        built_at=datetime.utcnow(),
    )
    db.add(version)

    html = b"<html><body>test</body></html>"
    blob = ArtifactBlob(
        id=str(uuid4()),
        version_id=version.id,
        blob_type="original",
        file_name="report.html",
        mime_type="text/html",
        file_size=len(html),
        checksum="x" * 64,
        data=html,
    )
    db.add(blob)

    artifact.current_version_id = version.id
    db.commit()
    db.refresh(artifact)
    return artifact


class TestExportService(unittest.TestCase):
    """ExportService.get_or_render + caching + eager-render logic."""

    def setUp(self):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        # In-memory SQLite.  The Artifact model has FKs, but the
        # export service only touches the artifacts table — no real
        # joins, so we don't need to create the rest of the schema.
        from app.models.artifact import (
            Artifact, ArtifactVersion, ArtifactBlob, MessageArtifact,
        )
        from sqlalchemy.ext.declarative import declarative_base
        # Manually create just the tables we need.
        engine = create_engine("sqlite:///:memory:")
        # Use the model's metadata; create_all will pick up the Artifact hierarchy.
        from app.models.base import Base
        Base.metadata.create_all(engine, tables=[
            Artifact.__table__,
            ArtifactVersion.__table__,
            ArtifactBlob.__table__,
            MessageArtifact.__table__,
        ])
        Session = sessionmaker(bind=engine)
        self.db = Session()
        self.artifact = _make_artifact_with_payload(self.db)

    def tearDown(self):
        self.db.close()

    def test_get_or_render_pdf_creates_blob(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        data, mime, file_name = svc.get_or_render(self.artifact, "pdf")
        self.assertTrue(data.startswith(b"%PDF-"))
        self.assertEqual(mime, "application/pdf")
        self.assertTrue(file_name.endswith(".pdf"))
        # Cached?
        available = svc.list_available_formats(self.artifact)
        self.assertIn("pdf", available)
        self.assertEqual(available["pdf"]["size"], len(data))

    def test_get_or_render_is_cached(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        data1, _, _ = svc.get_or_render(self.artifact, "xlsx")
        data2, _, _ = svc.get_or_render(self.artifact, "xlsx")
        # Same cached blob — exact bytes match
        self.assertEqual(data1, data2)

    def test_get_or_render_all_four_formats(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        for fmt in ("pdf", "pptx", "xlsx", "csv"):
            data, mime, file_name = svc.get_or_render(self.artifact, fmt)
            self.assertGreater(len(data), 50, f"{fmt} produced tiny output")
            self.assertIn(fmt, file_name.lower() or mime)

        available = svc.list_available_formats(self.artifact)
        self.assertEqual(set(available.keys()), {"pdf", "pptx", "xlsx", "csv"})

    def test_get_or_render_unknown_format_raises(self):
        from app.services.artifacts.exporters.service import ExportService, ExportError
        svc = ExportService(self.db)
        with self.assertRaises(ExportError):
            svc.get_or_render(self.artifact, "odp")

    def test_eager_render_default_user_signal_export(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        rendered = svc.eager_render_default(
            self.artifact,
            user_message="Export this as a PDF please",
            user_signal="export",
        )
        self.assertEqual(rendered, "pdf")
        self.assertIn("pdf", svc.list_available_formats(self.artifact))

    def test_eager_render_honors_pptx_hint(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        rendered = svc.eager_render_default(
            self.artifact,
            user_message="Make me a PowerPoint deck of this report",
            user_signal="export",
        )
        self.assertEqual(rendered, "pptx")

    def test_eager_render_honors_xlsx_hint(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        rendered = svc.eager_render_default(
            self.artifact,
            user_message="Save this as an Excel spreadsheet",
            user_signal="save",
        )
        self.assertEqual(rendered, "xlsx")

    def test_eager_render_noop_on_default_signal(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        rendered = svc.eager_render_default(
            self.artifact,
            user_message="Just looking at the report",
            user_signal="default",
        )
        self.assertIsNone(rendered)
        self.assertEqual(svc.list_available_formats(self.artifact), {})

    def test_eager_render_noop_when_already_cached(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        # Pre-render
        first = svc.eager_render_default(
            self.artifact, user_message="Export", user_signal="export"
        )
        self.assertEqual(first, "pdf")
        sizes_before = {k: v["size"] for k, v in svc.list_available_formats(self.artifact).items()}
        # Second call should be a no-op (still returns the format, doesn't re-render)
        second = svc.eager_render_default(
            self.artifact, user_message="Export", user_signal="export"
        )
        self.assertEqual(second, "pdf")
        sizes_after = {k: v["size"] for k, v in svc.list_available_formats(self.artifact).items()}
        self.assertEqual(sizes_before, sizes_after)

    def test_metadata_json_tracks_payload_formats(self):
        from app.services.artifacts.exporters.service import ExportService
        svc = ExportService(self.db)
        svc.get_or_render(self.artifact, "pdf")
        svc.get_or_render(self.artifact, "xlsx")
        self.db.refresh(self.artifact)
        meta = self.artifact.metadata_json or {}
        formats = meta.get("payload_formats") or {}
        self.assertIn("pdf", formats)
        self.assertIn("xlsx", formats)
        # Each entry has the expected fields
        for fmt_info in formats.values():
            self.assertIn("blob_id", fmt_info)
            self.assertIn("file_name", fmt_info)
            self.assertIn("mime_type", fmt_info)
            self.assertIn("size", fmt_info)


# ---------------------------------------------------------------------------
# Quick "smoke" — confirm the router can be imported and the new route
# is registered.  (Full HTTP test would need a test client + DB.)
# ---------------------------------------------------------------------------


class TestRouterRegistration(unittest.TestCase):
    def test_router_has_format_aware_routes(self):
        from app.routers.artifacts import router
        paths = sorted(r.path for r in router.routes if hasattr(r, "path"))
        # New routes registered?
        self.assertTrue(
            any("/download" in p for p in paths),
            "download route missing",
        )
        # Format-aware list route?
        self.assertTrue(
            any(p.endswith("/formats") for p in paths),
            "formats listing route missing",
        )


class TestChatPathPptxSmoke(unittest.TestCase):
    """Item 4: chat-path smoke test.

    Exercises the exact entry point the chat agent uses (FINALIZE →
    ``ExportService.get_or_render(artifact, "pptx", user_message=...)`` →
    ``_render_and_store`` → ``_render_deck_pipeline``) with all four PPT_*
    flags enabled, and asserts the deck renders, audits clean, and (when
    LibreOffice is available) thumbnails are stored.
    """

    FLAGS = {
        "PPT_DECK_PLANNER_ENABLED": True,
        "PPT_SMART_ROUTER_ENABLED": True,
        "PPT_LLM_POLISH_ENABLED": True,
        "PPT_AUDIT_ENABLED": True,
    }

    def _payload(self):
        return _sample_payload()

    def setUp(self):
        from app.config import settings
        self._saved = {}
        for flag, value in self.FLAGS.items():
            self._saved[flag] = getattr(settings, flag, False)
            setattr(settings, flag, value)
        # Use inline Postgres-style blob storage (no MinIO in CI/test env).
        self._saved["ARTIFACT_STORAGE_BACKEND"] = settings.ARTIFACT_STORAGE_BACKEND
        settings.ARTIFACT_STORAGE_BACKEND = "postgres_bytea"

    def tearDown(self):
        from app.config import settings
        for flag, value in self._saved.items():
            setattr(settings, flag, value)

    def test_chat_path_renders_pptx_with_all_flags(self):
        from app.services.artifacts.exporters.service import (
            BLOB_TYPE_FORMAT_EXPORT,
            ExportService,
        )
        from app.models.artifact import Artifact, ArtifactBlob, ArtifactVersion

        # In-memory SQLite DB mirroring the app's table set.
        import sqlalchemy as sa
        from app.database import Base
        engine = sa.create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        from sqlalchemy.orm import sessionmaker
        Session = sessionmaker(bind=engine)
        db = Session()

        artifact = _make_artifact_with_payload(db, self._payload())

        svc = ExportService(db)
        data, mime, ext = svc.get_or_render(
            artifact, "pptx",
            user_message="Build me a revenue deck",
            sql="SELECT month, revenue FROM sales",
        )
        self.assertTrue(ext.endswith(".pptx"), f"expected pptx filename, got {ext!r}")
        self.assertTrue(len(data) > 1000, "deck bytes suspiciously small")

        # Valid OOXML container with real slides (no placeholder-only deck).
        import zipfile, io, re
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            names = z.namelist()
            self.assertTrue(any(n.startswith("ppt/slides/slide") for n in names))
            self.assertTrue("ppt/presentation.xml" in names)
            # No "No chart data available" placeholder anywhere in the package.
            for n in names:
                if n.endswith(".xml"):
                    if b"No chart data available" in z.read(n):
                        self.fail(f"placeholder chart slide in {n}")

        # Audit report recorded on the artifact metadata (no FAIL).
        meta = artifact.metadata_json or {}
        reports = meta.get("audit_reports") or {}
        self.assertIn("pptx", reports, "expected audit report for pptx format")
        report = reports["pptx"]
        summary = report.get("summary") or {}
        self.assertEqual(summary.get("fail", 0), 0, f"audit FAILs: {summary}")
        self.assertGreater(summary.get("total", 0), 0, "audit ran no rules")
        self.assertIn(report.get("status"), ("PASS", "WARN"), report.get("status"))

        # Deck bytes persisted as an ArtifactBlob (source of truth for downloads).
        blob = (
            db.query(ArtifactBlob)
            .join(ArtifactVersion, ArtifactVersion.id == ArtifactBlob.version_id)
            .filter(
                ArtifactVersion.artifact_id == artifact.id,
                ArtifactBlob.blob_type == BLOB_TYPE_FORMAT_EXPORT,
            )
            .order_by(ArtifactBlob.created_date.desc())
            .first()
        )
        self.assertIsNotNone(blob, "deck blob not persisted")
        self.assertEqual(blob.blob_type, BLOB_TYPE_FORMAT_EXPORT)
        self.assertGreater(len(blob.data or b""), 1000, "stored blob too small")

        # Thumbnails stored for the deck when LibreOffice is available.
        thumbs = (
            db.query(ArtifactBlob)
            .join(ArtifactVersion, ArtifactVersion.id == ArtifactBlob.version_id)
            .filter(
                ArtifactVersion.artifact_id == artifact.id,
                ArtifactBlob.blob_type == "thumbnail",
            )
            .count()
        )
        self.assertGreater(thumbs, 0, "expected thumbnails for the deck")

    def test_chat_path_flags_default_off(self):
        """Global constraint check: all PPT_* flags default to False."""
        from app.config import Settings
        fields = Settings.model_fields if hasattr(Settings, "model_fields") else getattr(Settings, "__fields__", {})
        for flag in self.FLAGS:
            self.assertEqual(
                fields[flag].default, False,
                f"{flag} must default to False",
            )


if __name__ == "__main__":
    unittest.main(verbosity=2)
