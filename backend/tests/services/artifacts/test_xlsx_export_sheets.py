"""XLSX exporter tests — including the author-mode ``sheets`` path.

2026-08-29: the prompt instructs the agent to pass
``payload={"sheets": [...]}`` for Excel, but the tool allowlist rejected
``xlsx`` entirely (and the payload field was dropped). Both are fixed;
these tests pin the renderer's sheets behavior.
"""

from io import BytesIO

from openpyxl import load_workbook

from app.services.artifacts.exporters.xlsx_export import render
from app.services.synexia.contracts import ReportCardPayload


def _load(payload: dict) -> "Workbook":
    rcp = ReportCardPayload.model_validate(payload)
    data, mime, ext = render(rcp)
    assert ext == ".xlsx"
    assert mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert data
    return load_workbook(BytesIO(data))


def test_sheets_author_mode_renders_verbatim():
    payload = {
        "title": "Customers",
        "sheets": [
            {
                "title": "Customers",
                "rows": [
                    {"name": "ACME", "revenue": 1000},
                    {"name": "Globex", "revenue": 900},
                ],
            },
            {
                "title": "Regions",
                "columns": ["region", "sales"],
                "rows": [
                    {"region": "North", "sales": 500},
                    {"region": "South", "sales": 300},
                ],
                "summary": "Regional breakdown",
            },
        ],
    }
    wb = _load(payload)
    assert wb.sheetnames == ["Customers", "Regions"]

    ws = wb["Customers"]
    # Header row derived from row keys
    assert ws.cell(row=1, column=1).value == "name"
    assert ws.cell(row=1, column=2).value == "revenue"
    assert ws.cell(row=2, column=1).value == "ACME"
    assert ws.cell(row=3, column=2).value == 900

    ws2 = wb["Regions"]
    # Summary row first (merged), blank row, THEN the explicit columns
    assert ws2.cell(row=1, column=1).value == "Regional breakdown"
    assert ws2.cell(row=3, column=1).value == "region"
    assert ws2.cell(row=3, column=2).value == "sales"
    assert ws2.cell(row=4, column=2).value == 500   # North
    assert ws2.cell(row=5, column=2).value == 300   # South


def test_classic_layout_still_works_without_sheets():
    payload = {
        "title": "Sales",
        "summary": "July 2026",
        "kpis": [{"label": "Revenue", "value": "1,000"}],
    }
    wb = _load(payload)
    assert "Summary" in wb.sheetnames
    ws = wb["Summary"]
    assert ws.cell(row=1, column=1).value == "Sales"
    # The "Summary" heading exists somewhere in the classic layout
    found = any(
        ws.cell(row=r, column=1).value == "Summary"
        for r in range(1, 20)
    )
    assert found
