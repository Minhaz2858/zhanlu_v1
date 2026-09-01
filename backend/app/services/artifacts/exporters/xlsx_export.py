"""XLSX export — render a ReportCardPayload as a multi-sheet Excel workbook.

Uses openpyxl (MIT).  Sheet layout:

  1. **Summary** — title, source, generated-at, user_signal, summary,
     warnings, action chips
  2. **KPIs** — one row per KPI tile (label / value / delta / caption)
  3. **Data** — the chart's data, with auto-sized columns and a header
     style (only present if a chart has data)
  4. **Insights** — one row per insight bullet (icon, text)

This is a "spreadsheet view" of the same report — the user can slice /
pivot the data, or copy the KPI row into their own model.
"""

from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.synexia.contracts import ReportCardPayload
from app.services.artifacts.exporters._common import (
    ExportContext,
    chart_rows, chart_x_key, chart_y_keys, coerce_number,
    short_generated_at, insight_icon_to_emoji,
)


# --- Colors (ARGB hex strings, as openpyxl expects) --------------------------

C_PRIMARY = "FF2563EB"
C_TEXT = "FF0F172A"
C_MUTED = "FF64748B"
C_BORDER = "FFE2E8F0"
C_BG = "FFF8FAFC"
C_KPI_BG = "FFF1F5F9"
C_INSIGHT_BG = "FFEFF6FF"
C_WARN_BG = "FFFFFBEB"
C_WARN_BORDER = "FFF59E0B"
C_HEADER_BG = "FF1E293B"
C_WHITE = "FFFFFFFF"
C_DELTA_UP = "FF059669"
C_DELTA_DOWN = "FFDC2626"


_THIN_BORDER = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)


def render(payload: ReportCardPayload, ctx: Optional[ExportContext] = None) -> tuple[bytes, str, str]:
    ctx = ctx or ExportContext()
    wb = Workbook()
    # Replace the default sheet
    wb.remove(wb.active)

    if payload.sheets:
        # Author-mode: emit the LLM's sheets verbatim (one sheet per
        # logical view). Falls back to the classic layout when the agent
        # supplied none.
        _add_authored_sheets(wb, payload, ctx)
    else:
        _add_summary_sheet(wb, payload, ctx)
        if payload.kpis:
            _add_kpi_sheet(wb, payload)
        if payload.chart and payload.chart.data:
            _add_data_sheet(wb, payload)
        if payload.insights:
            _add_insights_sheet(wb, payload)

    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsx",
    )


def _add_authored_sheets(
    wb: Workbook, payload: ReportCardPayload, ctx: ExportContext
) -> None:
    """Render ``payload.sheets`` — the LLM's per-view spreadsheets.

    Each entry: {"title", "columns"?, "rows": [{col: value}], "summary"?}.
    Column order comes from ``columns`` when given, otherwise from the
    first row's keys (then union of all rows).
    """
    for idx, sheet in enumerate(payload.sheets):
        if not isinstance(sheet, dict):
            continue
        ws = wb.create_sheet(
            title=str(sheet.get("title") or f"Sheet {idx + 1}")[:31] or f"Sheet {idx + 1}",
            index=idx,
        )
        ws.sheet_view.showGridLines = False

        rows = sheet.get("rows") or []
        columns = sheet.get("columns")
        if not columns:
            cols_seen: list[str] = []
            for r in rows:
                if isinstance(r, dict):
                    for k in r.keys():
                        if k not in cols_seen:
                            cols_seen.append(k)
            columns = cols_seen

        summary = sheet.get("summary")
        row = 1
        if summary:
            c = ws.cell(row=row, column=1, value=str(summary))
            c.font = Font(name="Calibri", size=11, color=C_TEXT)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=max(len(columns), 1))
            ws.row_dimensions[row].height = 30
            row += 2

        if columns:
            for col_idx, col in enumerate(columns, start=1):
                c = ws.cell(row=row, column=col_idx, value=str(col))
                c.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
                c.fill = PatternFill(fill_type="solid", fgColor=C_HEADER_BG)
                c.border = _THIN_BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
            for r in rows:
                if not isinstance(r, dict):
                    continue
                for col_idx, col in enumerate(columns, start=1):
                    c = ws.cell(row=row, column=col_idx, value=r.get(col))
                    c.font = Font(name="Calibri", size=11, color=C_TEXT)
                    c.border = _THIN_BORDER
                row += 1
            for col_idx, col in enumerate(columns, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    12, min(48, 4 + 1.1 * len(str(col)))
                )


# --- Sheet builders ----------------------------------------------------------


def _add_summary_sheet(wb: Workbook, payload: ReportCardPayload, ctx: ExportContext) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    row = 1
    # Title — big, bold
    c = ws.cell(row=row, column=1, value=payload.title or "Zhanlu report")
    c.font = Font(name="Calibri", size=18, bold=True, color=C_TEXT)
    row += 1

    # Source
    if payload.source:
        c = ws.cell(row=row, column=1, value=f"Source: {payload.source}")
        c.font = Font(name="Calibri", size=11, color=C_MUTED)
        row += 1

    # Generated at
    c = ws.cell(row=row, column=1, value=f"Generated: {short_generated_at(payload)}")
    c.font = Font(name="Calibri", size=11, color=C_MUTED)
    row += 1

    # user_signal
    c = ws.cell(row=row, column=1, value=f"user_signal: {payload.user_signal or 'default'}")
    c.font = Font(name="Calibri", size=11, color=C_MUTED)
    row += 2  # blank line

    # Summary
    if payload.summary:
        c = ws.cell(row=row, column=1, value="Summary")
        c.font = Font(name="Calibri", size=14, bold=True, color=C_TEXT)
        row += 1
        c = ws.cell(row=row, column=1, value=payload.summary)
        c.font = Font(name="Calibri", size=11, color=C_TEXT)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60
        row += 2

    # Warnings
    if payload.warnings:
        c = ws.cell(row=row, column=1, value="Warnings")
        c.font = Font(name="Calibri", size=14, bold=True, color=C_WARN_BORDER)
        row += 1
        for w in payload.warnings:
            c = ws.cell(row=row, column=1, value=f"\u26A0  {w}")
            c.font = Font(name="Calibri", size=11, color=C_TEXT)
            c.fill = PatternFill("solid", fgColor=C_WARN_BG)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    # NOTE: payload.next_step is intentionally NOT rendered — it is
    # conversational guidance for the in-chat card, not workbook content.

    # Actions
    if payload.actions:
        c = ws.cell(row=row, column=1, value="Actions")
        c.font = Font(name="Calibri", size=14, bold=True, color=C_TEXT)
        row += 1
        for a in payload.actions:
            c = ws.cell(row=row, column=1, value=f"\u00BB {a.label}")
            c.font = Font(name="Calibri", size=11, bold=True, color=C_PRIMARY)
            row += 1
            c = ws.cell(row=row, column=1, value=a.prompt)
            c.font = Font(name="Calibri", size=10, color=C_MUTED)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 100


def _add_kpi_sheet(wb: Workbook, payload: ReportCardPayload) -> None:
    ws = wb.create_sheet("KPIs")
    ws.sheet_view.showGridLines = False

    # Header
    headers = ["Label", "Value", "Delta", "Caption"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _THIN_BORDER

    # Data
    for ri, k in enumerate(payload.kpis, start=2):
        cells = [
            (k.label or "", C_TEXT, False),
            (k.value or "—", C_TEXT, True),
            (k.delta or "", _delta_color(k.delta), True),
            (k.caption or "", C_MUTED, False),
        ]
        for ci, (val, color, bold) in enumerate(cells, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=11, color=color, bold=bold)
            c.fill = PatternFill("solid", fgColor=C_KPI_BG if ri % 2 == 0 else C_WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=(ci == 4))
            c.border = _THIN_BORDER

    # Column widths
    widths = [28, 20, 14, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze the header row
    ws.freeze_panes = "A2"


def _add_data_sheet(wb: Workbook, payload: ReportCardPayload) -> None:
    rows = chart_rows(payload)
    if not rows:
        return

    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False

    # Title row
    c = ws.cell(row=1, column=1, value=payload.chart.title or "Data")
    c.font = Font(name="Calibri", size=14, bold=True, color=C_TEXT)

    headers = list(rows[0].keys())
    header_row = 3
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _THIN_BORDER

    # Data rows
    for ri, r in enumerate(rows, start=header_row + 1):
        for ci, k in enumerate(headers, start=1):
            v = r.get(k, "")
            c = ws.cell(row=ri, column=ci, value=v)
            # Right-align numeric-looking values, left-align everything else
            if isinstance(v, (int, float)) or coerce_number(v) is not None:
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
            c.font = Font(name="Calibri", size=10, color=C_TEXT)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor=C_BG)
            c.border = _THIN_BORDER

    # Column widths — generous but capped
    for ci, h in enumerate(headers, start=1):
        col = get_column_letter(ci)
        max_len = len(str(h))
        for r in rows[:50]:
            v = r.get(h, "")
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 50)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _add_insights_sheet(wb: Workbook, payload: ReportCardPayload) -> None:
    ws = wb.create_sheet("Insights")
    ws.sheet_view.showGridLines = False

    # Header
    headers = ["Icon", "Text"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _THIN_BORDER

    for ri, ins in enumerate(payload.insights, start=2):
        c = ws.cell(row=ri, column=1, value=insight_icon_to_emoji(ins.icon))
        c.font = Font(name="Calibri", size=12, color=C_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _THIN_BORDER

        c = ws.cell(row=ri, column=2, value=ins.text)
        c.font = Font(name="Calibri", size=11, color=C_TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
        c.fill = PatternFill("solid", fgColor=C_INSIGHT_BG)
        c.border = _THIN_BORDER

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 100

    ws.freeze_panes = "A2"


# --- Helpers -----------------------------------------------------------------


def _delta_color(delta: Optional[str]) -> str:
    if not delta:
        return C_MUTED
    d = delta.strip()
    if d.startswith("+") or d.startswith("\u25B2"):
        return C_DELTA_UP
    if d.startswith("-") or d.startswith("\u25BC"):
        return C_DELTA_DOWN
    return C_MUTED
