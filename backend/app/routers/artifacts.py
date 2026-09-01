"""Artifact router — CRUD, preview, download, and lifecycle APIs.

All endpoints are permission-checked and scoped by org_id/app_id (inherited
from TimestampedBase).  Preview and download endpoints never expose raw file
paths — they stream blob data through the API with proper Content-Type.
"""

import io
import logging
import re
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query, Response, Depends, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.services.artifacts.artifact_service import ArtifactService
from app.services.artifacts.exporters import SUPPORTED_FORMATS
from app.services.artifacts.exporters._theme import list_themes
from app.services.artifacts.exporters.service import ExportError, ExportService
from app.config import settings
from app.services.artifacts.event_logger import log_deck_event_fire_and_forget
from app.services.artifacts.preview_builder import (
    convert_to_preview,
    generate_thumbnail,
    extract_docx_outline,
    extract_pptx_outline,
    build_ms_word_open_url,
    convert_docx_to_html,
    convert_pptx_to_html,
)

logger = logging.getLogger(__name__)


def _log_deck_download(artifact, user_id: Optional[str], fmt: str) -> None:
    """Phase 5 — fire-and-forget deck_downloaded event (pptx only)."""
    try:
        log_deck_event_fire_and_forget(
            None,
            "deck_downloaded",
            artifact_id=getattr(artifact, "id", None),
            user_id=user_id,
            metadata={"format": fmt},
            org_id=getattr(artifact, "org_id", "default-org"),
            app_id=getattr(artifact, "app_id", "default-app"),
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("deck download event log failed: %s", exc)


router = APIRouter(tags=["artifacts"])


# Characters disallowed in the legacy ``filename=`` parameter of
# ``Content-Disposition`` (RFC 6266 / RFC 5987 say the legacy form must
# be ASCII-quoted — bytes 0x20-0x7E excluding quote/backslash).  Any
# character outside this set must be sent in the ``filename*=UTF-8''…``
# form.  We use this together with :func:`content_disposition_header`
# to keep file names with non-ASCII characters (em-dash, accented
# letters, emoji, …) from blowing up the response with a
# ``UnicodeEncodeError`` raised when Starlette writes the header.
_CONTENT_DISPOSITION_SAFE = re.compile(r"[ -~!#-&(-\[\]-~]*")  # matches safe bytes


def _content_disposition(disposition: str, file_name: str) -> str:
    """Build a ``Content-Disposition`` value that supports Unicode names.

    Always emits a UTF-8 ``filename*=`` parameter so titles like
    ``Sales_Report_—_Address_Distribution_by_Region.html`` round-trip
    correctly.  The legacy ``filename=`` form is included only when
    ``file_name`` is pure ASCII; otherwise it's replaced with an
    underscore-prefixed safe fallback so old browsers still see a
    non-empty filename.
    """
    safe_ascii = file_name.encode("ascii", "replace").decode("ascii").replace("?", "_")
    # Strip control chars and quotes from the ASCII fallback
    safe_ascii = re.sub(r'[\x00-\x1f\x7f"\\]', "_", safe_ascii) or "download"
    encoded = quote(file_name, safe="")
    return f"{disposition}; filename=\"{safe_ascii}\"; filename*=UTF-8''{encoded}"


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# --- Request/Response schemas ---

class CreateArtifactRequest(BaseModel):
    artifact_type: str = Field(..., description="pptx | docx | pdf | md | html | chart | dashboard | mini_app | image | xlsx")
    title: str
    description: Optional[str] = None
    conversation_id: Optional[str] = None
    execution_id: Optional[str] = None
    created_by_agent_id: Optional[str] = None
    data_snapshot_ids: Optional[list[str]] = None


class StoreBlobRequest(BaseModel):
    """Base64-encoded blob data for direct storage (used by skills)."""
    blob_type: str = Field("original", description="original | preview | thumbnail")
    file_name: str
    mime_type: str
    data_base64: str


class UpdateStatusRequest(BaseModel):
    status: str


class LinkToMessageRequest(BaseModel):
    message_id: str
    conversation_id: str
    display_order: int = 0


class CreateVersionRequest(BaseModel):
    changelog: Optional[str] = None
    source_json: Optional[dict] = None
    produced_by_skill: Optional[str] = None
    sandbox_job_id: Optional[str] = None


# --- Endpoints ---

@router.get("/artifacts")
def list_artifacts(
    conversation_id: Optional[str] = Query(None),
    artifact_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(50, le=200),
    offset: int = Query(0),
    db: Session = Depends(get_db),
):
    """List artifacts with optional filters."""
    service = ArtifactService(db)
    artifacts = service.list_artifacts(
        conversation_id=conversation_id,
        artifact_type=artifact_type,
        status=status,
        limit=limit,
        offset=offset,
    )
    return [a.to_dict() for a in artifacts]


@router.post("/artifacts")
def create_artifact(req: CreateArtifactRequest, db: Session = Depends(get_db)):
    """Create a new artifact (starts in 'draft' status)."""
    service = ArtifactService(db)
    try:
        artifact = service.create_artifact(
            artifact_type=req.artifact_type,
            title=req.title,
            description=req.description,
            conversation_id=req.conversation_id,
            execution_id=req.execution_id,
            created_by_agent_id=req.created_by_agent_id,
            data_snapshot_ids=req.data_snapshot_ids,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return artifact.to_dict()


@router.get("/artifacts/{artifact_id}")
def get_artifact(artifact_id: str, db: Session = Depends(get_db)):
    """Get artifact detail including versions, preview modes, and outline."""
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    result = artifact.to_dict()
    result["versions"] = [v.to_dict() for v in service.get_versions(artifact_id)]

    # Document-specific preview metadata (Task 4 DOCX, Task 5 PPTX)
    original = service.get_original_blob(artifact_id)
    if artifact.artifact_type == "docx":
        if original is not None and original.data:
            result["preview_outline"] = extract_docx_outline(original.data)
        else:
            result["preview_outline"] = []

        modes: list[str] = ["pdf", "self_hosted_html"]
        public_url = (settings.APP_PUBLIC_URL or "").rstrip("/")
        if public_url and original and original.data:
            result["ms_word_open_url"] = build_ms_word_open_url(
                public_url=public_url,
                artifact_id=artifact_id,
                file_name=original.file_name or f"{artifact.title or artifact_id}.docx",
            )
            modes.insert(0, "ms_word")
        else:
            result["ms_word_open_url"] = None
        result["preview_modes"] = modes
    elif artifact.artifact_type == "pptx":
        if original is not None and original.data:
            result["preview_outline"] = extract_pptx_outline(original.data)
        else:
            result["preview_outline"] = []
        # PPTX has no MS Word Online equivalent. The faithful preview
        # is the LibreOffice→PDF path ("pdf" mode); "self_hosted_html"
        # is the positioned-slide HTML renderer.
        result["preview_modes"] = ["pdf", "self_hosted_html"]
        result["ms_word_open_url"] = None
    elif artifact.artifact_type == "xlsx":
        # XLSX: faithful preview via LibreOffice→PDF. No inline HTML
        # renderer exists yet, so "pdf" is the only preview mode.
        result["preview_outline"] = []
        result["preview_modes"] = ["pdf"]
        result["ms_word_open_url"] = None
    else:
        result["preview_modes"] = []
        result["preview_outline"] = []
        result["ms_word_open_url"] = None

    return result


@router.patch("/artifacts/{artifact_id}/status")
def update_artifact_status(artifact_id: str, req: UpdateStatusRequest, db: Session = Depends(get_db)):
    """Update artifact lifecycle status."""
    service = ArtifactService(db)
    try:
        artifact = service.update_status(artifact_id, req.status)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")
    return artifact.to_dict()


# --- Version endpoints ---

@router.get("/artifacts/{artifact_id}/versions")
def list_versions(artifact_id: str, db: Session = Depends(get_db)):
    """List all versions of an artifact."""
    service = ArtifactService(db)
    versions = service.get_versions(artifact_id)
    return [v.to_dict() for v in versions]


@router.post("/artifacts/{artifact_id}/versions")
def create_version(artifact_id: str, req: CreateVersionRequest, db: Session = Depends(get_db)):
    """Create a new version of an artifact."""
    service = ArtifactService(db)
    version = service.create_version(
        artifact_id=artifact_id,
        changelog=req.changelog,
        source_json=req.source_json,
        produced_by_skill=req.produced_by_skill,
        sandbox_job_id=req.sandbox_job_id,
    )
    if not version:
        raise HTTPException(status_code=404, detail="Artifact not found")
    return version.to_dict()


# ── Interactive canvas: save a user-edited HTML snapshot as a new version ──

class CanvasSaveRequest(BaseModel):
    """Body for the interactive-canvas edit loop.

    ``html`` is the full edited HTML document the user produced in the
    canvas editor.  Saving creates an immutable new version (source_json
    carries the HTML), stores the original + preview blobs, and marks the
    version built — so the chat preview and download both reflect the
    edit without an LLM round-trip.
    """
    html: str = Field(..., min_length=1, description="Full edited HTML document")
    changelog: Optional[str] = Field(default="Edited in canvas")
    source: str = Field(default="user", description="user | llm")


@router.post("/artifacts/{artifact_id}/canvas/save", response_model=dict)
def canvas_save(
    artifact_id: str,
    req: CanvasSaveRequest,
    db: Session = Depends(get_db),
):
    """Save user-edited HTML from the interactive canvas as a new version.

    Pipeline (mirrors the LLM create_artifact flow, minus the LLM call):
      1. create_version with source_json={"html": ..., "source": req.source}
      2. store original blob (text/html, UTF-8)
      3. store preview blob (same HTML — self-contained preview)
      4. mark_version_built
    Returns {artifact_id, version_id, version_number, preview_url}.
    """
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    html_bytes = req.html.encode("utf-8")
    version = service.create_version(
        artifact_id=artifact_id,
        changelog=req.changelog,
        source_json={"html": req.html, "source": req.source},
        produced_by_skill=None,
    )
    if not version:
        raise HTTPException(status_code=404, detail="Artifact not found")

    # Original blob (text/html)
    service.store_blob(
        version_id=version.id,
        blob_type="original",
        file_name=f"{artifact.title or 'artifact'}.html",
        mime_type="text/html",
        data=html_bytes,
    )
    # Preview blob — same self-contained HTML for the inline iframe.
    service.store_blob(
        version_id=version.id,
        blob_type="preview",
        file_name=f"{artifact.title or 'artifact'}.html",
        mime_type="text/html",
        data=html_bytes,
    )
    service.mark_version_built(version.id)

    return {
        "success": True,
        "artifact_id": artifact_id,
        "version_id": version.id,
        "version_number": version.version_number,
        "source": req.source,
        "preview_url": f"/api/artifacts/{artifact_id}/preview",
        "download_url": f"/api/artifacts/{artifact_id}/download",
    }


@router.post("/artifacts/{artifact_id}/versions/{version_id}/blobs")
def store_blob(
    artifact_id: str,
    version_id: str,
    req: StoreBlobRequest,
    db: Session = Depends(get_db),
):
    """Store a blob (original/preview/thumbnail) for a version.

    Accepts base64-encoded data.  Used by skills to upload generated files.
    If blob_type is 'original' and no preview exists, auto-generates preview.
    """
    import base64

    service = ArtifactService(db)

    try:
        data = base64.b64decode(req.data_base64)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 data")

    blob = service.store_blob(
        version_id=version_id,
        blob_type=req.blob_type,
        file_name=req.file_name,
        mime_type=req.mime_type,
        data=data,
    )

    # Auto-generate preview if storing original and no preview exists
    if req.blob_type == "original":
        artifact = service.get_artifact(artifact_id)
        if artifact:
            preview_result = convert_to_preview(data, req.file_name, artifact.artifact_type)
            if preview_result:
                preview_data, preview_name, preview_mime = preview_result
                service.store_blob(
                    version_id=version_id,
                    blob_type="preview",
                    file_name=preview_name,
                    mime_type=preview_mime,
                    data=preview_data,
                )

                # Try to generate thumbnail from PDF preview
                if preview_mime == "application/pdf":
                    thumb = generate_thumbnail(preview_data)
                    if thumb:
                        service.store_blob(
                            version_id=version_id,
                            blob_type="thumbnail",
                            file_name=f"{req.file_name.rsplit('.', 1)[0]}_thumb.png",
                            mime_type="image/png",
                            data=thumb,
                        )

            # Mark version as built (preview_ready)
            service.mark_version_built(version_id)

    return blob.to_dict()


# --- Preview / Download endpoints ---

@router.get("/artifacts/{artifact_id}/preview")
def get_preview(
    artifact_id: str,
    format: Optional[str] = Query(
        None,
        description="If set, returns the requested preview variant. "
                    "Supported: 'pdf' (PDF render), 'html' (DOCX → sanitized HTML).",
    ),
    db: Session = Depends(get_db),
):
    """Get the preview blob (PDF/image/HTML) for inline display.

    Streams the binary data with proper Content-Type.  Never exposes file paths.

    If ``?format=pdf`` is provided, renders the PDF on demand and
    serves it inline.  This is the "Open in new tab" path the
    ReportCard UI uses.

    If ``?format=html`` is provided, converts a stored DOCX to
    sanitized HTML via mammoth and returns it directly.
    """
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    fmt = (format or "").lower().strip() if format else ""

    if fmt == "pdf":
        # ── Faithful PDF for native office-file artifacts ──
        # For docx/pptx/xlsx artifacts the *original blob IS the file the
        # user downloads*.  Converting that exact blob to PDF via
        # LibreOffice guarantees "what you see is what you get" — the
        # preview matches the download byte-for-byte.  This is different
        # from html_report artifacts, where the PDF is re-rendered from
        # the ReportCardPayload source data (a deliberately different
        # rendering pipeline).
        if artifact.artifact_type in ("docx", "pptx", "xlsx"):
            # 1. Try the cached PDF preview blob first (instant).
            cached_blob = service.get_preview_blob(artifact_id)
            if cached_blob and cached_blob.mime_type == "application/pdf":
                cached_data = service.get_blob_data(cached_blob)
                if cached_data:
                    return Response(
                        content=cached_data,
                        media_type="application/pdf",
                        headers={
                            "Content-Disposition": _content_disposition(
                                "inline", cached_blob.file_name or f"{artifact_id}.pdf"
                            ),
                            "Cache-Control": "public, max-age=3600",
                        },
                    )

            # 2. No cached PDF — convert the actual original file on
            #    demand.  convert_to_preview uses LibreOffice headless
            #    (serialized by _LIBREOFFICE_LOCK) and returns None if
            #    LibreOffice is unavailable or the conversion fails.
            original = service.get_original_blob(artifact_id)
            original_data = service.get_blob_data(original) if original else None
            if original_data:
                preview_result = convert_to_preview(
                    original_data,
                    original.file_name or f"{artifact_id}.{artifact.artifact_type}",
                    artifact.artifact_type,
                )
                if preview_result:
                    pdf_data, pdf_name, pdf_mime = preview_result
                    # Cache as a preview blob so the next request is instant.
                    version = service.get_current_version(artifact_id)
                    if version:
                        try:
                            service.store_blob(
                                version_id=version.id,
                                blob_type="preview",
                                file_name=pdf_name,
                                mime_type=pdf_mime,
                                data=pdf_data,
                            )
                        except Exception as cache_err:
                            logger.warning(
                                "Failed to cache PDF preview for %s: %s",
                                artifact_id, cache_err,
                            )
                    return Response(
                        content=pdf_data,
                        media_type=pdf_mime,
                        headers={
                            "Content-Disposition": _content_disposition("inline", pdf_name),
                            "Cache-Control": "public, max-age=3600",
                        },
                    )
            # 3. Conversion failed (e.g. LibreOffice not installed, or
            #    no original blob).  Fall through to the ExportService
            #    path below as a last resort — it re-renders from the
            #    ReportCardPayload, which is less faithful but better
            #    than a 404.
            logger.info(
                "Faithful PDF conversion unavailable for %s artifact %s — "
                "falling back to ExportService",
                artifact.artifact_type, artifact_id,
            )

        # ── ExportService path (html_report → PDF, and fallback) ──
        exporter = ExportService(db)
        try:
            data, mime, file_name = exporter.get_or_render(artifact, "pdf")
        except ExportError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            logger.exception("PDF preview render failed for %s", artifact_id)
            raise HTTPException(status_code=500, detail=f"Preview failed: {e}")
        return Response(
            content=data,
            media_type=mime,
            headers={
                "Content-Disposition": _content_disposition("inline", file_name),
                "Cache-Control": "public, max-age=3600",
            },
        )

    if fmt == "html":
        if artifact.artifact_type not in ("docx", "pptx"):
            raise HTTPException(
                status_code=400,
                detail="?format=html is only supported for DOCX and PPTX artifacts",
            )
        # Robust blob lookup. The "original" blob is the canonical source
        # for DOCX/PPTX, but the sandbox / finalize paths may store the
        # file under a different blob_type (e.g. "preview") or the
        # artifact may only have a derived render. Fall back through
        # preview blob, then any blob on the current version whose mime
        # is a real docx/pptx file. This guarantees the inline chat
        # preview works regardless of how the artifact was produced.
        target_mimes = {
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        }
        def _has_bytes(b):
            # A blob is usable if it has inline data OR a storage_uri
            # that the storage backend can resolve. The ORM ``blob.data``
            # attribute is only populated for legacy inline blobs;
            # modern blobs store bytes behind ``storage_uri``
            # (e.g. ``inline://<id>``) and require ``get_blob_data``
            # to load. Without this guard the preview would treat
            # every storage-backed blob as "empty" and fall back.
            return b is not None and (b.data is not None or b.storage_uri)

        def _mime_acceptable(b):
            """Return True if the blob's stored mime matches the docx/pptx
            target list OR can be reasonably inferred to be a real
            docx/pptx file by its extension.

            We accept the generic ``application/octet-stream`` (and an
            empty mime) for ``.docx`` / ``.pptx`` files because some
            container images register mime types incompletely — see
            ``container_manager._FALLBACK_MIME_BY_EXT``.  Without this
            tolerance, perfectly valid Office files get filtered out
            and the preview falls back to "Inline preview is not
            available" even though mammoth / pptx2html can read them.
            """
            if not b:
                return False
            if b.mime_type and b.mime_type in target_mimes:
                return True
            ext = (b.file_name or "").rsplit(".", 1)[-1].lower() if b.file_name else ""
            if ext in ("docx", "pptx") and (
                not b.mime_type
                or b.mime_type == "application/octet-stream"
                or b.mime_type == "application/vnd.ms-office"
            ):
                return True
            return False

        blob = service.get_original_blob(artifact_id)
        if (not _has_bytes(blob) or not _mime_acceptable(blob)):
            blob = service.get_preview_blob(artifact_id)
        if (not _has_bytes(blob) or not _mime_acceptable(blob)):
            current = service.get_current_version(artifact_id)
            if current:
                for candidate in service.get_version_blobs(current.id):
                    if (candidate and _has_bytes(candidate)
                            and _mime_acceptable(candidate)):
                        blob = candidate
                        break
        # Resolve the actual bytes via the storage backend. This is
        # the canonical way to read blob content (handles both inline
        # data and storage_uri-backed blobs). The download endpoint
        # uses the same resolver; the preview must too, otherwise
        # storage-backed blobs render the fallback even though the
        # file is a perfectly valid .docx on disk.
        blob_bytes = service.get_blob_data(blob) if blob else None
        if not blob or not blob_bytes:
            # Graceful fallback: return a clean HTML fragment with a
            # download link so the inline chat preview never breaks with
            # a raw 404. The frontend renders this as the document body.
            safe_title = (artifact.title or "Document").replace("<", "&lt;")
            ext = artifact.artifact_type or "docx"
            fallback = (
                '<div style="padding:32px 24px;font-family:system-ui,-apple-system,Segoe UI,sans-serif;color:#374151;">'
                f'<h3 style="margin:0 0 8px;font-size:16px;">{safe_title}</h3>'
                '<p style="margin:0 0 16px;color:#6b7280;font-size:14px;">'
                "Inline preview is not available for this file. "
                "Please download it to view the contents."
                "</p>"
                f'<a href="/api/artifacts/{artifact_id}/download" '
                'style="display:inline-block;padding:8px 16px;background:#2563eb;'
                'color:#fff;border-radius:6px;text-decoration:none;font-size:14px;">'
                f"Download .{ext}</a></div>"
            )
            return Response(
                content=fallback,
                media_type="text/html; charset=utf-8",
                headers={"Cache-Control": "no-store"},
            )
        from app.services.artifacts.preview_builder import (
            convert_docx_to_html, convert_pptx_to_html,
        )
        if artifact.artifact_type == "docx":
            html, _messages = convert_docx_to_html(blob_bytes)
            kind = "DOCX"
        else:  # pptx
            html, _messages = convert_pptx_to_html(blob_bytes)
            kind = "PPTX"
        if not html:
            # Conversion failed (e.g. mammoth can't read a malformed
            # file). Return a clean fallback rather than 500 so the
            # preview panel still shows something useful.
            safe_title = (artifact.title or "Document").replace("<", "&lt;")
            ext = artifact.artifact_type or "docx"
            fallback = (
                '<div style="padding:32px 24px;font-family:system-ui,-apple-system,Segoe UI,sans-serif;color:#374151;">'
                f'<h3 style="margin:0 0 8px;font-size:16px;">{safe_title}</h3>'
                '<p style="margin:0 0 16px;color:#6b7280;font-size:14px;">'
                f"Inline {kind} preview is temporarily unavailable. "
                "Please download the file to view it."
                "</p>"
                f'<a href="/api/artifacts/{artifact_id}/download" '
                'style="display:inline-block;padding:8px 16px;background:#2563eb;'
                'color:#fff;border-radius:6px;text-decoration:none;font-size:14px;">'
                f"Download .{ext}</a></div>"
            )
            return Response(
                content=fallback,
                media_type="text/html; charset=utf-8",
                headers={"Cache-Control": "no-store"},
            )
        return Response(
            content=html,
            media_type="text/html; charset=utf-8",
            headers={
                "Content-Disposition": _content_disposition("inline", f"{artifact_id}.html"),
                "Cache-Control": "public, max-age=60",
            },
        )

    # Default: return whatever preview blob is already stored. For
    # HTML-native artifacts (html / html_report) the original blob IS the
    # preview — fall back to it so inline preview works even when no
    # derived preview blob was ever generated (e.g. finalize_into_artifact
    # stores the report HTML as the original blob only).
    blob = service.get_preview_blob(artifact_id)
    if not blob and artifact.artifact_type in ("html", "html_report"):
        blob = service.get_original_blob(artifact_id)
    if not blob:
        raise HTTPException(status_code=404, detail="Preview not available")
    blob_data = service.get_blob_data(blob)
    if blob_data is None:
        raise HTTPException(status_code=404, detail="Preview data not available")
    return Response(
        content=blob_data,
        media_type=blob.mime_type,
        headers={
            "Content-Disposition": _content_disposition("inline", blob.file_name or "preview"),
            "Cache-Control": "public, max-age=3600",
        },
    )


@router.get("/artifacts/{artifact_id}/formats")
def list_artifact_formats(artifact_id: str, db: Session = Depends(get_db)):
    """List the export formats that are already cached on this artifact.

    Returns ``{"formats": {"pdf": {...}, "pptx": {...}, ...}}`` — each
    entry has the file_name, mime_type, size, and blob_id for the
    cached render.  An empty dict means no exports have been rendered
    yet (the frontend should expect to render-on-demand on download).
    """
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    exporter = ExportService(db)
    formats = exporter.list_available_formats(artifact)
    return {"artifact_id": artifact_id, "formats": formats}


@router.get("/artifacts/themes")
def list_deck_themes():
    """Return the catalog of selectable deck themes for the export picker.

    Each entry carries a 4-color swatch so the frontend can render theme
    chips without loading the full token files.  Used by the PPTX export
    action surface (Phase 1 design-engine).
    """
    return {"themes": list_themes()}


@router.post("/artifacts/templates/analyze")
async def analyze_template_endpoint(file: UploadFile = File(...)):
    """Analyze an uploaded ``.pptx`` template.

    Extracts the slide-layout + placeholder ``idx/type/inches`` map (the
    ``ppt-template-creator`` methodology) — the first step toward brand-
    template-driven rendering.  Returns the structured layout map.
    """
    from app.services.artifacts.exporters._advanced import analyze_template
    data = await file.read()
    if not data[:4] == b"PK\x03\x04":
        raise HTTPException(
            status_code=400, detail="Uploaded file is not a valid PPTX (ZIP).")
    try:
        return analyze_template(data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Template analysis failed: {e}")


@router.get("/artifacts/{artifact_id}/download")
def download_artifact(
    artifact_id: str,
    format: Optional[str] = Query(
        None,
        description="Optional output format. One of: pdf, pptx, xlsx, csv. "
                    "If omitted, returns the artifact's original blob. "
                    "If specified, renders (and caches) the report in that format.",
    ),
    theme: Optional[str] = Query(
        None,
        description="Deck theme name (e.g. ocean-depths, sunset-boulevard). "
                    "Only affects pptx. See GET /artifacts/themes. "
                    "When set, the render is fresh (not cached).",
    ),
    mode: Optional[str] = Query(None, description="light (default) or dark."),
    style_recipe: Optional[str] = Query(
        None,
        description="Corner-radius/spacing recipe: sharp (default), soft, rounded, pill.",
    ),
    doc_type: Optional[str] = Query(
        None,
        description=(
            "Document-type awareness (DOCX only). One of: report (default — "
            "cover+TOC+body), brief (compact, no cover/TOC), memo (To/From/Date/Subject "
            "header). When set to anything other than report, the render is fresh "
            "(not cached)."
        ),
    ),
    force: bool = Query(
        False,
        description="Re-render the artifact from its source payload instead of "
                    "returning a cached format blob. Use after a render-code fix "
                    "or to discard a stale cached document.",
    ),
    db: Session = Depends(get_db),
):
    """Download the artifact file.

    Without ``?format=`` returns the original blob (the HTML report
    stored by FINALIZE).  With ``?format=pdf|pptx|xlsx|csv`` returns
    a freshly-rendered export in that format — the bytes are cached on
    the artifact's current version, so repeat downloads are instant.

    The PDF path is also reachable via the preview endpoint, so the
    frontend can open the file in a new tab without a download
    prompt.
    """
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    fmt = (format or "").lower().strip() if format else ""

    if not fmt:
        # No format requested — return the original blob.
        blob = service.get_original_blob(artifact_id)
        if not blob:
            raise HTTPException(status_code=404, detail="Artifact file not available")
        blob_data = service.get_blob_data(blob)
        if blob_data is None:
            raise HTTPException(status_code=404, detail="Artifact data not available")
        if getattr(artifact, "artifact_type", None) == "pptx":
            _log_deck_download(artifact, user_id=None, fmt="original")
        return Response(
            content=blob_data,
            media_type=blob.mime_type,
            headers={
                "Content-Disposition": _content_disposition("attachment", blob.file_name or "download"),
            },
        )

    if fmt not in SUPPORTED_FORMATS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported format {format!r}. Supported: {', '.join(SUPPORTED_FORMATS)}",
        )

    # Format requested — look up the cached render or generate one.
    exporter = ExportService(db)
    try:
        data, mime, file_name = exporter.get_or_render(
            artifact, fmt,
            theme=theme, mode=mode, style_recipe=style_recipe, doc_type=doc_type,
            force=force,
        )
    except ExportError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("Export render failed for %s format=%s", artifact_id, fmt)
        raise HTTPException(status_code=500, detail=f"Export failed: {e}")

    if getattr(artifact, "artifact_type", None) == "pptx":
        _log_deck_download(artifact, user_id=None, fmt=fmt)
    return Response(
        content=data,
        media_type=mime,
        headers={
            "Content-Disposition": _content_disposition("attachment", file_name or "download"),
        },
    )


@router.get("/artifacts/{artifact_id}/thumbnail")
def get_thumbnail(artifact_id: str, db: Session = Depends(get_db)):
    """Get the thumbnail image for an artifact (used in artifact cards)."""
    service = ArtifactService(db)
    version = service.get_current_version(artifact_id)
    if not version:
        raise HTTPException(status_code=404, detail="No version available")
    blobs = service.get_version_blobs(version.id, blob_type="thumbnail")
    if not blobs:
        raise HTTPException(status_code=404, detail="Thumbnail not available")
    blob = blobs[0]
    blob_data = service.get_blob_data(blob)
    if blob_data is None:
        raise HTTPException(status_code=404, detail="Thumbnail data not available")
    return Response(
        content=blob_data,
        media_type=blob.mime_type,
        headers={"Cache-Control": "public, max-age=3600"},
    )


# --- Message linking ---

@router.post("/artifacts/{artifact_id}/link")
def link_to_message(artifact_id: str, req: LinkToMessageRequest, db: Session = Depends(get_db)):
    """Link an artifact to a chat message for inline preview display."""
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")
    link = service.link_to_message(
        artifact_id=artifact_id,
        message_id=req.message_id,
        conversation_id=req.conversation_id,
        display_order=req.display_order,
    )
    return link.to_dict()


# --- File-format export endpoint ---------------------------------------------

class ExportArtifactRequest(BaseModel):
    """Body for the one-click export endpoint.

    The ReportCard export menu hits this when the user wants a specific
    file format (docx / pptx / xlsx / pdf / md) and the html_report
    artifact already exists.
    """
    rows: list[dict] = Field(
        default_factory=list,
        description="Data rows to include in the export. Typically the "
                    "chart.data from a prior ReportCardPayload.",
    )
    title: str = Field(..., description="Title for the generated file.")
    instructions: Optional[str] = Field(
        None,
        description="Natural-language instructions passed to the sandbox "
                    "skill (e.g. 'Include KPIs and a chart section').",
    )
    conversation_id: Optional[str] = None
    agent_app_id: Optional[str] = None


@router.post("/artifacts/{artifact_id}/export/{format}")
def export_artifact(
    artifact_id: str,
    format: str,
    req: ExportArtifactRequest,
    db: Session = Depends(get_db),
):
    """Render a downloadable file for an existing artifact.

    Re-uses the existing ``run_sandbox_skill`` pipeline so the file is
    produced in an isolated Docker sandbox, stored as a new Artifact
    (linked as a sibling), and rendered to PDF for inline preview.

    The newly-created artifact's id is returned so the frontend can
    mount ``ArtifactPreviewCard`` for it.
    """
    fmt = (format or "").lower().strip()
    if fmt not in {"docx", "pptx", "xlsx", "pdf", "html", "md"}:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported format {format!r}. "
                   "Supported: docx, pptx, xlsx, pdf, html, md.",
        )

    # Confirm the source artifact exists so we can 404 cleanly.
    service = ArtifactService(db)
    source = service.get_artifact(artifact_id)
    if not source:
        raise HTTPException(status_code=404, detail="Source artifact not found")

    if not req.rows:
        raise HTTPException(
            status_code=400,
            detail="rows must be a non-empty array of objects",
        )

    # Lazy import to keep module-load time light.
    from app.services.tool_handlers.sandbox_tool import run_sandbox_skill_sync

    try:
        result = run_sandbox_skill_sync(
            args={
                "format": fmt,
                "data": req.rows,
                "title": req.title,
                "instructions": req.instructions or "",
            },
            db=db,
            user_id=req.agent_app_id,
            context={
                "conversation_id": req.conversation_id,
                "agent_app_id": req.agent_app_id,
            },
        )
    except Exception as e:
        logger.exception("export_artifact failed for %s format=%s", artifact_id, fmt)
        raise HTTPException(status_code=500, detail=f"Export failed: {e}")

    if not result.get("success"):
        raise HTTPException(
            status_code=400,
            detail=result.get("error") or "Sandbox skill failed",
        )

    return {
        "source_artifact_id": artifact_id,
        "artifact_id": result.get("artifact_id"),
        "format": fmt,
        "title": result.get("title"),
        "preview_url": result.get("preview_url"),
        "download_url": result.get("download_url"),
        "job_id": result.get("job_id"),
        "files": result.get("files", []),
    }


@router.get("/messages/{message_id}/artifacts")
def get_message_artifacts(message_id: str, db: Session = Depends(get_db)):
    """Get all artifacts linked to a message (for inline preview rendering)."""
    service = ArtifactService(db)
    return service.get_message_artifacts(message_id)


# --- Round-trip editing (P1.3) ----------------------------------------------


class EditArtifactRequest(BaseModel):
    """Targeted edit operations applied to a stored PPTX/DOCX blob.

    The edit is atomic: either every operation applies and a new version
    is created, or the request fails with 400 and the artifact is
    untouched.  See ``artifacts/editors/`` for the operation shapes.
    """

    format: str = Field(..., description="pptx | docx — which stored blob to edit")
    operations: list[dict] = Field(..., description="non-empty list of edit ops")
    changelog: Optional[str] = None


@router.post("/artifacts/{artifact_id}/edit")
def edit_artifact(
    artifact_id: str,
    req: EditArtifactRequest,
    db: Session = Depends(get_db),
):
    """Apply targeted edits to an artifact's PPTX/DOCX blob (round-trip).

    "Change the title on slide 3" without a full regeneration: reads the
    current blob, applies the operations, stores the result as a new
    immutable ``ArtifactVersion``, re-runs the semantic audit, and marks
    the version built.  Cached format exports live on the old version, so
    downloads naturally re-render against the edited content.
    """
    from app.services.artifacts.editors import EditError, apply_edits

    fmt = (req.format or "").lower().strip()
    if fmt not in ("pptx", "docx"):
        raise HTTPException(status_code=400, detail="format must be pptx or docx")

    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    # Locate the source blob: native pptx/docx artifacts edit their
    # original blob; html_report artifacts edit the cached format export.
    blob = None
    if artifact.artifact_type == fmt:
        blob = service.get_original_blob(artifact_id)
    if blob is None:
        version = service.get_current_version(artifact_id)
        if version:
            for b in service.get_version_blobs(version.id):
                if (b.file_name or "").lower().endswith(f".{fmt}"):
                    blob = b
                    break
    if blob is None:
        raise HTTPException(
            status_code=400,
            detail=f"No editable {fmt} blob found on this artifact",
        )
    data = service.get_blob_data(blob)
    if not data:
        raise HTTPException(status_code=400, detail="Blob data unavailable")

    try:
        edited, applied = apply_edits(fmt, data, req.operations)
    except EditError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # New immutable version for the edit.
    changelog = req.changelog or "; ".join(applied)
    version = service.create_version(
        artifact_id=artifact_id,
        changelog=changelog,
        produced_by_skill="round-trip-edit",
    )
    if not version:
        raise HTTPException(status_code=500, detail="Failed to create version")

    new_blob = service.store_blob(
        version_id=version.id,
        blob_type="original" if artifact.artifact_type == fmt else "format_export",
        file_name=blob.file_name or f"edited.{fmt}",
        mime_type=blob.mime_type,
        data=edited,
    )

    # Re-run the semantic audit on the edited file (P0 loop).
    audit_report = ExportService(db)._run_semantic_audit(fmt, edited)
    if audit_report:
        meta = dict(artifact.metadata_json or {})
        audits = dict(meta.get("audit_reports") or {})
        audits[fmt] = audit_report
        meta["audit_reports"] = audits
        artifact.metadata_json = meta
        db.commit()

    service.mark_version_built(version.id, validation_report=audit_report)

    return {
        "artifact_id": artifact_id,
        "version_id": version.id,
        "version_number": version.version_number,
        "blob_id": new_blob.id,
        "applied": applied,
        "audit_status": (audit_report or {}).get("status"),
    }


# --- Outline gate (P0.2) -----------------------------------------------------


@router.post("/artifacts/{artifact_id}/outline/approve")
def approve_outline(
    artifact_id: str,
    user_message: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Approve a pending outline (ZHANLU_OUTLINE_GATE=block deployments).

    Flips ``metadata_json["outline"]["approved"]``, moves the artifact
    from ``draft`` to ``preview_ready``, and kicks off the deferred
    eager render so the first download is instant.  Idempotent —
    approving an already-approved outline is a no-op.
    """
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")

    meta = dict(artifact.metadata_json or {})
    outline = dict(meta.get("outline") or {})
    if not outline:
        raise HTTPException(status_code=400, detail="Artifact has no outline to approve")

    already = bool(outline.get("approved"))
    if not already:
        outline["approved"] = True
        meta["outline"] = outline
        artifact.metadata_json = meta
        if artifact.status == "draft":
            artifact.status = "preview_ready"
        db.commit()

    # Kick the deferred eager render (best-effort).
    rendered = None
    try:
        exporter = ExportService(db)
        rendered = exporter.eager_render_default(
            artifact,
            user_message=user_message or meta.get("user_message") or "",
            user_signal=meta.get("user_signal", "default"),
            sql=meta.get("sql"),
            source=meta.get("source"),
            org_id=artifact.org_id,
            app_id=artifact.app_id,
        )
    except Exception as e:
        logger.warning("outline approve: eager render failed for %s: %s", artifact_id, e)

    return {
        "artifact_id": artifact_id,
        "approved": True,
        "was_already_approved": already,
        "status": artifact.status,
        "eager_rendered": rendered,
    }


@router.get("/artifacts/{artifact_id}/outline")
def get_outline(artifact_id: str, db: Session = Depends(get_db)):
    """Return the artifact's outline spec (P0.2), or 404 when absent."""
    service = ArtifactService(db)
    artifact = service.get_artifact(artifact_id)
    if not artifact:
        raise HTTPException(status_code=404, detail="Artifact not found")
    outline = (artifact.metadata_json or {}).get("outline")
    if not outline:
        raise HTTPException(status_code=404, detail="No outline recorded for this artifact")
    return outline


# --- Brand kit (P1.2, multi-tenant) ------------------------------------------


class BrandKitRequest(BaseModel):
    """The workspace brand kit document.  All fields optional; at least
    one valid color is required.  Unknown keys are ignored."""

    name: Optional[str] = None
    colors: Optional[dict] = None
    fonts: Optional[dict] = None
    logo_blob_uri: Optional[str] = None


@router.get("/workspaces/brand-kit")
def get_workspace_brand_kit(
    org_id: str = Query("default-org"),
    app_id: str = Query("default-app"),
    db: Session = Depends(get_db),
):
    """Return the workspace's brand kit (or ``{"brand_kit": null}``)."""
    from app.services.artifacts.brand_kit import get_brand_kit

    return {"brand_kit": get_brand_kit(db, org_id=org_id, app_id=app_id)}


@router.put("/workspaces/brand-kit")
def put_workspace_brand_kit(
    req: BrandKitRequest,
    org_id: str = Query("default-org"),
    app_id: str = Query("default-app"),
    db: Session = Depends(get_db),
):
    """Create or replace the workspace's brand kit.

    Once set, every artifact export (PPTX / DOCX / HTML report) in this
    workspace renders in the brand palette by default, and previously
    cached brand renders are invalidated automatically (fingerprint
    mismatch on the next download).
    """
    from app.services.artifacts.brand_kit import set_brand_kit

    try:
        kit = set_brand_kit(
            db,
            req.model_dump(exclude_none=True),
            org_id=org_id,
            app_id=app_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"brand_kit": kit}


@router.delete("/workspaces/brand-kit")
def delete_workspace_brand_kit(
    org_id: str = Query("default-org"),
    app_id: str = Query("default-app"),
    db: Session = Depends(get_db),
):
    """Remove the workspace's brand kit (exports revert to zhanlu-blue)."""
    from app.services.artifacts.brand_kit import clear_brand_kit

    removed = clear_brand_kit(db, org_id=org_id, app_id=app_id)
    return {"removed": removed}


@router.post("/workspaces/brand-kit/palette")
async def extract_brand_palette(file: UploadFile = File(...)):
    """Extract a suggested brand palette from an uploaded logo image.

    Returns a ``colors`` dict the client can review and then persist via
    ``PUT /workspaces/brand-kit``.  Nothing is stored by this endpoint.
    """
    from app.services.artifacts.brand_kit import extract_palette_from_image

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty upload")
    try:
        colors = extract_palette_from_image(data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"colors": colors}


# ── Generic file preview by URL ────────────────────────────────────
#
# Used by the frontend FilePreviewer for files that point at our own
# /api/uploads/ directory.  The endpoint:
#   • accepts only local /api/uploads/ URLs (no SSRF, no proxying
#     arbitrary external URLs)
#   • resolves the URL to settings.upload_path / <safe filename>
#   • converts DOCX/PPTX/XLSX/MD to a sanitized HTML document that
#     browsers can render inline
#   • passes through HTML, PDF, images, audio, video as their original
#     bytes with the correct Content-Type
#
# This is what makes "preview a generated .docx report from
# /my-files" actually work in the browser — the backend serves an
# HTML rendition of the Office file.  No LibreOffice required for
# DOCX/PPTX/XLSX → HTML (mammoth + python-pptx + openpyxl all do it
# in-process).
_PREVIEW_PREVIEW_WRAP = """<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{title}</title>
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; max-width: 960px; margin: 32px auto; padding: 24px; line-height: 1.6; color: #1a1a1a; background: #fff; }}
h1, h2, h3, h4 {{ color: #111; margin-top: 1.4em; }}
h1 {{ border-bottom: 2px solid #e5e5e5; padding-bottom: 6px; }}
table {{ border-collapse: collapse; width: 100%; margin: 12px 0; }}
th, td {{ border: 1px solid #d4d4d4; padding: 6px 10px; text-align: left; }}
th {{ background: #f5f5f5; }}
pre, code {{ background: #f6f6f6; padding: 2px 4px; border-radius: 3px; }}
pre {{ padding: 12px; overflow-x: auto; }}
.slide {{ border: 1px solid #e5e5e5; border-radius: 8px; padding: 24px; margin: 16px 0; min-height: 200px; background: #fafafa; }}
.sheet {{ margin: 16px 0; }}
.sheet-title {{ font-weight: 600; margin: 8px 0; }}
</style></head>
<body>
{body}
</body></html>"""


def _wrap_preview(body: str, title: str) -> bytes:
    return _PREVIEW_PREVIEW_WRAP.format(title=title, body=body).encode("utf-8")


def _convert_xlsx_to_html(xlsx_bytes: bytes) -> str:
    """Render an XLSX as a sequence of HTML tables (one per sheet).

    Uses openpyxl to read each sheet, but no styling — just a faithful
    row-by-row dump.  Heavy enough for inline preview of typical
    agent-generated reports (a few hundred rows per sheet).
    """
    try:
        from openpyxl import load_workbook  # lazy
    except ImportError:
        return ""

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("openpyxl failed to open xlsx: %s", exc)
        return ""

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f'<h2 class="sheet-title">{sheet_name}</h2>')
        parts.append('<table>')
        for row in ws.iter_rows(values_only=True):
            cells = "".join(
                "<td>{}</td>".format(
                    ("" if v is None else str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
                )
                for v in row
            )
            parts.append(f"<tr>{cells}</tr>")
        parts.append("</table>")
    wb.close()
    return "".join(parts) if parts else ""


@router.get("/files/preview")
def preview_file_by_url(url: str = Query(..., description="Local /api/uploads/ path")):
    """Convert a file (by /api/uploads/ URL) to a previewable format.

    Returns a sanitized HTML document for Office files (DOCX/PPTX/XLSX)
    and Markdown, or the original bytes for HTML/PDF/images/audio/video.
    """
    from urllib.parse import urlparse

    # ── Security: only accept local /api/uploads/ paths ──
    raw = url
    if raw.startswith("http://") or raw.startswith("https://"):
        parsed = urlparse(raw)
        raw = parsed.path
    if not raw.startswith("/api/uploads/"):
        raise HTTPException(
            status_code=400,
            detail="Only local /api/uploads/ URLs are previewable",
        )

    filename = raw[len("/api/uploads/"):]
    # Disallow path traversal / nested paths
    if not filename or "/" in filename or ".." in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid file path")

    file_path = settings.upload_path / filename
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found on disk")

    ext = file_path.suffix.lower().lstrip(".")
    display_name = filename
    raw_bytes = file_path.read_bytes()

    if ext == "docx":
        html, _ = convert_docx_to_html(raw_bytes)
        if not html:
            raise HTTPException(status_code=500, detail="DOCX conversion failed")
        return Response(content=_wrap_preview(html, display_name), media_type="text/html")

    if ext == "pptx":
        html, _ = convert_pptx_to_html(raw_bytes)
        if not html:
            raise HTTPException(status_code=500, detail="PPTX conversion failed")
        return Response(content=_wrap_preview(html, display_name), media_type="text/html")

    if ext == "xlsx":
        html = _convert_xlsx_to_html(raw_bytes)
        if not html:
            raise HTTPException(status_code=500, detail="XLSX conversion failed")
        return Response(content=_wrap_preview(html, display_name), media_type="text/html")

    if ext in ("md", "markdown"):
        try:
            import markdown
            body = markdown.markdown(
                raw_bytes.decode("utf-8", errors="replace"),
                extensions=["tables", "fenced_code", "codehilite", "toc"],
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="markdown package not installed")
        return Response(content=_wrap_preview(body, display_name), media_type="text/html")

    if ext in ("html", "htm"):
        return Response(content=raw_bytes, media_type="text/html")

    if ext == "pdf":
        return Response(content=raw_bytes, media_type="application/pdf")

    # Pass-through binary types
    pass_through_mime = {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "webp": "image/webp",
        "svg": "image/svg+xml",
        "bmp": "image/bmp",
        "mp3": "audio/mpeg",
        "m4a": "audio/mp4",
        "wav": "audio/wav",
        "ogg": "audio/ogg",
        "flac": "audio/flac",
        "aac": "audio/aac",
        "mp4": "video/mp4",
        "mov": "video/quicktime",
        "webm": "video/webm",
        "ogv": "video/ogg",
        "mkv": "video/x-matroska",
        "txt": "text/plain",
        "csv": "text/csv",
        "json": "application/json",
        "xml": "application/xml",
    }
    if ext in pass_through_mime:
        return Response(content=raw_bytes, media_type=pass_through_mime[ext])

    raise HTTPException(
        status_code=415,
        detail=f"Preview not supported for .{ext} files",
    )
